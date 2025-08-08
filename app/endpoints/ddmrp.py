from fastapi import APIRouter, HTTPException
from typing import Dict, Any, List
import sys
import os
import pandas as pd
import math
import json
import openpyxl


# Add the parent directories to sys.path
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(root_dir)

# Import the existing function that reads from processed data
from logic.ddmrp_engine import calculate_ddmrp_plan

router = APIRouter()

OUTPUTS_DIR = "data/outputs"
PROCESSED_DIR = "data/processed"


def get_latest_analysis_data():
    """
    Get the latest analysis data from either outputs or processed directory.
    """
    try:
        # Try CSV output first
        output_path = os.path.join(OUTPUTS_DIR, "ddmrp_weekly_production_plan.csv")
        if os.path.exists(output_path):
            return pd.read_csv(output_path)

        # Fall back to processed directory
        processed_path = os.path.join(PROCESSED_DIR, "latest_analysis.xlsx")
        if os.path.exists(processed_path):
            return pd.read_excel(processed_path)

        return None
    except Exception as e:
        print(f"Error loading analysis data: {e}")
        return None


def get_available_skus_list():
    """Get list of available SKUs from processed data (renamed to avoid conflict)"""
    try:
        output_path = os.path.join(OUTPUTS_DIR, "ddmrp_weekly_production_plan.csv")
        if os.path.exists(output_path):
            df = pd.read_csv(output_path)
            return df['SKU'].unique().tolist()
        return []
    except:
        return []


@router.get("/sku/{sku_id}")
async def get_sku_data(sku_id: str) -> Dict[str, Any]:
    """
    Get DDMRP data for a specific SKU from processed data.
    This fixes the original 'list' object has no attribute 'get' error.
    """
    try:
        # Call the existing function that reads from processed data
        result = calculate_ddmrp_plan(sku_id)

        if "error" in result:
            available_skus = get_available_skus_list()
            raise HTTPException(
                status_code=404,
                detail=f"SKU {sku_id} not found. Available SKUs: {available_skus}"
            )

        return result

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving SKU data: {str(e)}")


@router.get("/dashboard")
async def get_dashboard_data() -> Dict[str, List[Dict[str, Any]]]:
    """
    Returns summary data for all SKUs from processed data.
    This fixes the original dashboard error by using real data instead of hardcoded SKUs.
    """
    try:
        # Get list of available SKUs from processed data
        available_skus = get_available_skus_list()

        if not available_skus:
            return {"skus": []}

        dashboard_data = []

        for sku_id in available_skus:
            try:
                # Call the existing function for each SKU
                sku_data = calculate_ddmrp_plan(str(sku_id))

                if sku_data and "error" not in sku_data:
                    # Handle different field name variations
                    net_flow = sku_data.get("net_flow", sku_data.get("Net Flow", 0))
                    red_zone = sku_data.get("red_zone", sku_data.get("Red Zone", 0))
                    yellow_zone = sku_data.get("yellow_zone", sku_data.get("Yellow Zone", 0))
                    weekly_adu = sku_data.get("weekly_adu", sku_data.get("Weekly ADU", 0))
                    current_inventory = sku_data.get("inventory_(current_week)", sku_data.get("Inventory (Current Week)", 0))
                    recommended_production = sku_data.get("recommended_production", sku_data.get("Recommended Production", 0))
                    target_week = sku_data.get("target_production_week", sku_data.get("Target Production Week", ""))

                    # Determine status based on buffer zones
                    if net_flow <= red_zone:
                        status = "red"
                    elif net_flow <= red_zone + yellow_zone:
                        status = "yellow"
                    else:
                        status = "green"

                    dashboard_item = {
                        "sku": str(sku_id),
                        "status": status,
                        # Add all CSV columns with proper field name handling
                        "current_week": sku_data.get("current_week", sku_data.get("Current Week", "")),
                        "target_production_week": sku_data.get("target_production_week",
                                                               sku_data.get("Target Production Week", "")),
                        "weekly_adu": sku_data.get("weekly_adu", sku_data.get("Weekly ADU", 0)),
                        "cov": sku_data.get("cov", sku_data.get("CoV", 0)),
                        "red_base": sku_data.get("red_base", sku_data.get("Red Base", 0)),
                        "red_safety": sku_data.get("red_safety", sku_data.get("Red Safety", 0)),
                        "red_zone": sku_data.get("red_zone", sku_data.get("Red Zone", 0)),
                        "yellow_zone": sku_data.get("yellow_zone", sku_data.get("Yellow Zone", 0)),
                        "green_zone": sku_data.get("green_zone", sku_data.get("Green Zone", 0)),
                        "top_of_red": sku_data.get("top_of_red", sku_data.get("Top of Red", 0)),
                        "top_of_yellow": sku_data.get("top_of_yellow", sku_data.get("Top of Yellow", 0)),
                        "top_of_green": sku_data.get("top_of_green", sku_data.get("Top of Green", 0)),
                        "net_flow": net_flow,
                        "recommended_production": sku_data.get("recommended_production",
                                                               sku_data.get("Recommended Production", 0)),
                        "inventory_(current_week)": sku_data.get("inventory_(current_week)",
                                                                 sku_data.get("Inventory (Current Week)", 0)),
                        "on_order": sku_data.get("on_order", sku_data.get("On Order", 0)),
                        "qualified_demand": sku_data.get("qualified_demand", sku_data.get("Qualified Demand", 0)),

                        # Keep legacy field names for backward compatibility
                        "netFlow": net_flow,
                        "weeklyADU": sku_data.get("weekly_adu", sku_data.get("Weekly ADU", 0)),
                        "currentInventory": sku_data.get("inventory_(current_week)",
                                                         sku_data.get("Inventory (Current Week)", 0)),
                        "recommendedProduction": sku_data.get("recommended_production",
                                                              sku_data.get("Recommended Production", 0)),
                        "targetWeek": sku_data.get("target_production_week", sku_data.get("Target Production Week", ""))
                    }

                    dashboard_data.append(dashboard_item)

            except Exception as e:
                print(f"Error processing SKU {sku_id}: {e}")
                continue

        return {"skus": dashboard_data}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving dashboard data: {str(e)}")


@router.get("/dashboard/status/{status}")
async def get_dashboard_by_status(status: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Returns dashboard data filtered by status (red, yellow, green)
    """
    if status not in ["red", "yellow", "green"]:
        raise HTTPException(status_code=400, detail="Status must be 'red', 'yellow', or 'green'")

    # Get all dashboard data
    all_data = await get_dashboard_data()

    # Filter by status
    filtered_data = [item for item in all_data["skus"] if item["status"] == status]

    return {"skus": filtered_data}


@router.get("/available-skus")
async def get_available_skus_endpoint() -> Dict[str, List[str]]:
    """
    Returns list of all available SKUs from processed data
    """
    try:
        skus = get_available_skus_list()
        return {"skus": [str(sku) for sku in skus]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving available SKUs: {str(e)}")


@router.get("/summary")
async def get_analysis_summary() -> Dict[str, Any]:
    """
    Get summary statistics of the current analysis
    """
    try:
        df = get_latest_analysis_data()

        if df is None:
            return {"status": "no_data", "message": "No analysis data available"}

        # Handle different column name variations
        net_flow_col = 'net_flow' if 'net_flow' in df.columns else 'Net Flow'
        red_zone_col = 'red_zone' if 'red_zone' in df.columns else 'Red Zone'
        yellow_zone_col = 'yellow_zone' if 'yellow_zone' in df.columns else 'Yellow Zone'
        recommended_production_col = 'recommended_production' if 'recommended_production' in df.columns else 'Recommended Production'

        # Calculate summary statistics
        total_skus = len(df)
        red_count = len(df[df[net_flow_col] <= df[red_zone_col]])
        yellow_count = len(
            df[(df[net_flow_col] > df[red_zone_col]) & (df[net_flow_col] <= df[red_zone_col] + df[yellow_zone_col])])
        green_count = len(df[df[net_flow_col] > df[red_zone_col] + df[yellow_zone_col]])

        total_recommended_production = df[recommended_production_col].sum()
        average_net_flow = df[net_flow_col].mean()

        return {
            "status": "available",
            "total_skus": total_skus,
            "status_counts": {
                "red": red_count,
                "yellow": yellow_count,
                "green": green_count
            },
            "total_recommended_production": float(total_recommended_production),
            "average_net_flow": float(average_net_flow),
            "last_updated": os.path.getmtime(os.path.join(PROCESSED_DIR, "latest_analysis.xlsx")) if os.path.exists(
                os.path.join(PROCESSED_DIR, "latest_analysis.xlsx")) else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving analysis summary: {str(e)}")


@router.get("/ddmrp/sku-details/{sku_id}")
def get_sku_details(sku_id: str):
    file_path = f"data/outputs/SKU_{sku_id}.xlsx"

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail={
            "error": "SKU not found",
            "message": f"No analysis data available for SKU {sku_id}"
        })

    try:
        df = pd.read_excel(file_path, sheet_name="Master Data")

        if df.empty:
            raise ValueError("Master Data sheet is empty")

        first_row = df.iloc[0]

        def safe_int(value, default=0):
            return int(value) if pd.notna(value) and not math.isnan(value) else default

        def safe_float(value, default=0.0):
            return float(value) if pd.notna(value) and not math.isnan(value) else default

        def safe_str(value, default="N/A"):
            return str(value) if pd.notna(value) else default

        result = {
            "sku_id": safe_str(first_row.get("Product ID")),
            "Product Description": safe_str(first_row.get("Product Desc")),
            "MOQ": safe_int(first_row.get("MOQ")),
            "ADU": safe_float(first_row.get("ADU")),
            "Lead Time": safe_int(first_row.get("Lead Time")),
            "MRP Type": safe_str(first_row.get("MRP Type")),
            "Red Zone": safe_float(first_row.get("Red Zone")),
            "Yellow Zone": safe_float(first_row.get("Yellow Zone")),
            "Green Zone": safe_float(first_row.get("Green Zone")),
            "historical": []
        }

        for _, row in df.iterrows():
            week = row.get("Week")
            if pd.notna(week):
                result["historical"].append({
                    "week": safe_str(week),
                    "inventory": safe_int(row.get("Inventory")),
                    "quantitySold": safe_int(row.get("Quantity Sold")),
                    "productionOrders": safe_int(row.get("Production Orders"))
                })

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail={
            "error": "Internal Server Error",
            "message": str(e)
        })


@router.get("/artikel-materialien")
async def get_artikel_materialien() -> List[Dict[str, Any]]:
    """
    Return the contents of the 'Artikel & Materialien FGR+' sheet as JSON.
    """
    file_path = os.path.join("data", "inputs", "Artikel & Materialien FGR+.XLSX")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Artikel & Materialien file not found")
    df = pd.read_excel(file_path, sheet_name="Artikel FGR+", skiprows=1)
    return df.to_dict(orient="records")

@router.post("/artikel-materialien")
async def update_artikel_materialien(records: List[Dict[str, Any]]):
    """
    Accept a list of row dictionaries and overwrite the 'Artikel FGR+' tab with the new data.
    """
    file_path = os.path.join("data", "inputs", "Artikel & Materialien FGR+.XLSX")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Artikel & Materialien file not found")
    try:
        # Load the entire workbook so that other sheets are preserved
        book = openpyxl.load_workbook(file_path)
        # Remove the old sheet if it exists
        if "Artikel FGR+" in book.sheetnames:
            idx = book.sheetnames.index("Artikel FGR+")
            book.remove(book.worksheets[idx])
        # Create a DataFrame from posted data and write it as the new sheet
        df_new = pd.DataFrame(records)
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            writer.book = book
            df_new.to_excel(writer, sheet_name="Artikel FGR+", index=False)
        return {"status": "success", "rows": len(df_new)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to update file: {e}")